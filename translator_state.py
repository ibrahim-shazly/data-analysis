from selenium import webdriver
from selenium.webdriver import ChromeOptions, Chrome
from selenium.webdriver.common.by import By
from openpyxl import workbook, load_workbook
from openpyxl.styles import PatternFill
import time
from selenium.webdriver.support.ui import WebDriverWait
from urllib.parse import unquote

# 1- أفتح صفحة المنتجات على المتصفح
opts = ChromeOptions()
# opts.add_experimental_option("detach", True)

browser = Chrome(options=opts)
browser.implicitly_wait(0.5)

# 2- فتح ملف الإكسيل
wb = load_workbook('translator_state.xlsx')
ws = wb['روايات']


# لوب الشغلانة
for row in range(191, 191 + 1):
    # تحضير الرابط من الإكسيل
    n_s_link = ws["F" + str(row)].value
    n_link = "https://kolnovel.com/series" + str(n_s_link)

    # الدخول للرابط على المتصفح
    browser.get(n_link)
    time.sleep(1.5)

    # التأكد من إن الرابط الجديد متطابق مع القديم

    current_url = browser.current_url
    if (n_link+"/") == current_url:
        # لو تطابق الرابط القديم والجديد
        print("links are the same")
    else:
        # أو لم يتطابقا
        print("NOOOOOOOOOOOO, links are the same")
        print(n_link+"/")
        print(unquote(str(current_url)))
        # غير الرابط القديم
        ws["F" + str(row)].value = unquote(str(current_url)).replace("https://kolnovel.com/series","")[:-1]
        # غير لون الخلية للأصفر 
        ws["F" + str(row)].fill = PatternFill(fill_type="solid",start_color="FFFF00",end_color="FFFF00")


    # متغير النهارده شهر كام ###################
    thisyear = "2022"
    thismonth = "أكتوبر"
    newmonth = "نوفمبر"

    lastchapterdate = browser.find_elements(By.CSS_SELECTOR, "div.epl-date")

    print("last date is" + lastchapterdate[0].text)
    print(len(lastchapterdate))

    findthisyear = lastchapterdate[0].text.find(thisyear)
    findthismonth = lastchapterdate[0].text.find(thismonth)
    findnewmonth = lastchapterdate[0].text.find(newmonth)

    if findthisyear > -1 and findthismonth > -1:
        # شوف عدد الفصول هذا الشهر
        ch_n = 0

        for n in range(0 , len(lastchapterdate)):
            date_ch_txt =  lastchapterdate[n].get_attribute('textContent')
            print(str(n) + " - " + date_ch_txt)
            # لو لقيت فصل في الشهر الحالي
            if date_ch_txt.find(thisyear)  > -1 and date_ch_txt.find(thismonth) > -1:
                ch_n = ch_n + 1
            else:
                break

        if ch_n > 0:
            # حط في الإكسيل الرواية مستمرة
            ws["A" + str(row)].value = ws["A4"].value
            # حط في الإكسيل عدد الفصول
            ws["G" + str(row)].value = ch_n

            browser.execute_script("window.scrollTo(0, 230)")
            time.sleep(1)
            browser.find_element(By.CSS_SELECTOR, "div.lastend > div:nth-child(2) > a").click()
            time.sleep(2)

            ch_text = browser.find_element(By.CSS_SELECTOR, "#kol_content").text
            word_counter = len(ch_text.split())
            print(str(word_counter) + " word")
            ws["H" + str(row)].value = word_counter


            wb.save('translator_state.xlsx')
            continue


        for date_ch in lastchapterdate:
            date_ch_txt =  date_ch.get_attribute('textContent')
            print(date_ch_txt)
            if findthisyear == -1 or findthismonth == -1:
                print("preaking chapters counter loop")
                break
            if thismonth in date_ch_txt:
                ch_n = ch_n + 1
                print("there is " + str(ch_n))
        # ضع عدد الفصول في الإكسيل
        ws["G" + str(row)].value = ch_n
    
    elif findthisyear  > -1 and findnewmonth > -1:
        n = 0
        ch_n = 0
        for n in range(0 , len(lastchapterdate) + 1):
            date_ch_txt =  lastchapterdate[n].get_attribute('textContent')
            print(str(n) + " - " + date_ch_txt)

            # لو لقيت فصل في الشهر الحالي
            if date_ch_txt.find(thisyear)  > -1 and date_ch_txt.find(thismonth) > -1:
                ch_n = ch_n + 1

            elif date_ch_txt.find(thisyear)  > -1 and date_ch_txt.find(newmonth) > -1:
                None
            else:
                break

        if ch_n > 0:
            # حط في الإكسيل الرواية مستمرة
            ws["A" + str(row)].value = ws["A4"].value
            # حط في الإكسيل عدد الفصول
            ws["G" + str(row)].value = ch_n

    else:
        print("novel is !!!!!not ongoing")
        ws["A" + str(row)].value = ws["A5"].value
    browser.execute_script("window.scrollTo(0, 230)")
    time.sleep(1)
    browser.find_element(By.CSS_SELECTOR, "div.ts-chl-collapsible-content > div > ul > li:nth-child(1) > a").click()
    time.sleep(2)

    ch_text = browser.find_element(By.CSS_SELECTOR, "#kol_content").text
    word_counter = len(ch_text.split())
    print(str(word_counter) + " word")
    ws["H" + str(row)].value = word_counter


    wb.save('translator_state.xlsx')


###### أنا حاليا بحاول أخليه يجيب عددا الفصول الشهرية

    # # إدخال الأرباح
    # user_rev_unformatted = ws["H" + str(row)].value
    # user_rev = "{:.2f}".format(user_rev_unformatted)
    # print(user_rev)
    # browser.find_element(By.CSS_SELECTOR, "body > div > div > section.content > div:nth-child(4) > div > div.box-body > form > input.form-control").click()
    # time.sleep(0.75)
    # browser.find_element(By.CSS_SELECTOR, "body > div > div > section.content > div:nth-child(4) > div > div.box-body > form > input.form-control").send_keys(user_rev)
    # time.sleep(0.75)

    # browser.find_element(By.CSS_SELECTOR, "body > div > div > section.content > div:nth-child(4) > div > div.box-body > form > button").click()
    
    # # تحويل إلى مستحقات متراكمة
    # browser.find_element(By.CSS_SELECTOR, "body > div > div > section.content > div:nth-child(11) > div > div.box-body > form > button").click()