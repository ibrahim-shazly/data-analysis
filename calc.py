from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import (
    DateRange,
    Dimension,
    Metric,
    RunReportRequest,
    Filter,
    FilterExpression,
)
from openpyxl import workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill


# غتح ملف الإكسيل
wb = load_workbook("rev.xlsx")
ws = wb["روايات 70% (NEW)"]
# e_row = 10
for e_row in range(10, 217 + 1):
    # جلب الرابط
    n_link = ws["L" + str(e_row)].value

    print(n_link)

    def sample_run_report(
        property_id="313712184", credentials_json_path="client_secrets.json"
    ):
        """Runs a simple report on a Google Analytics 4 property."""
        # TODO(developer): Uncomment this variable and replace with your
        #  Google Analytics 4 property ID before running the sample.
        property_id = "313712184"

        # [START analyticsdata_json_credentials_initialize]
        # TODO(developer): Uncomment this variable and replace with a valid path to
        #  the credentials.json file for your service account downloaded from the
        #  Cloud Console.
        credentials_json_path = "client_secrets.json"

        # Explicitly use service account credentials by specifying
        # the private key file.
        client = BetaAnalyticsDataClient.from_service_account_json(
            credentials_json_path
        )
        # [END analyticsdata_json_credentials_initialize]

        # [START analyticsdata_json_credentials_run_report]
        request = RunReportRequest(
            property=f"properties/{property_id}",
            dimensions=[Dimension(name="year")],
            metrics=[Metric(name="screenPageViews")],
            date_ranges=[DateRange(start_date="2024-02-01", end_date="2024-02-29")],
            dimension_filter=FilterExpression(
                filter=Filter(
                    field_name="pageReferrer",
                    string_filter=Filter.StringFilter(
                        match_type="CONTAINS", value=n_link
                    ),
                )
            ),
        )
        response = client.run_report(request)
        # [END analyticsdata_json_credentials_run_report]

        print("Report result:")
        for row in response.rows:
            print(row.dimension_values[0].value, row.metric_values[0].value)
        global n_views

        try:
            n_views = row.metric_values[0].value
        except:
            n_views = 0

    # [END analyticsdata_json_credentials_quickstart]

    if __name__ == "__main__":
        sample_run_report()

    try:
        print(n_views)
    except:
        n_views = 0
        print(n_views)

    # وضع ربح الرواية في الخلية
    ws["C" + str(e_row)].value = int(n_views)

    value = 0
    n_views = 0
    # حفظ ملف الإكسيل

    wb.save("rev.xlsx")
