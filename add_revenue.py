from selenium import webdriver
from selenium.webdriver import ChromeOptions, Chrome
from selenium.webdriver.common.by import By
from openpyxl import workbook, load_workbook
import time
from selenium.webdriver.support.ui import WebDriverWait

# 1- أفتح صفحة المنتجات على المتصفح
opts = ChromeOptions()
# opts.add_experimental_option("detach", True)

browser = Chrome(options=opts)
browser.implicitly_wait(0.5)

# 2- فتح ملف الإكسيل
wb = load_workbook("rev.xlsx")
ws = wb["إضافة الأرباح"]
num = -1

# 3- تسجيل الدخول
browser.get("https://teams2.kolnovel.com/login")
time.sleep(2)
# إدخال الإيميل
browser.find_element(By.CSS_SELECTOR, "#email").click()
time.sleep(0.75)
browser.find_element(By.CSS_SELECTOR, "#email").send_keys("me.hima.1221@gmail.com")
time.sleep(0.75)
# إدخال الباسبورد
browser.find_element(By.CSS_SELECTOR, "#password").click()
time.sleep(0.75)
browser.find_element(By.CSS_SELECTOR, "#password").send_keys("???????????")
time.sleep(0.75)
# الضغط على زر تسجيل الدخول
browser.find_element(
    By.CSS_SELECTOR,
    "#app > main > div > div > div > div > div.card-body > form > div.form-group.row.mb-0 > div > button",
).click()
time.sleep(1.5)

# أنت الأن قد دخلت الموقع

# تعريف المتغيرات
for row in range(2, 69 + 1):
    user_num = ws["K" + str(row)].value
    user_link = "https://teams2.kolnovel.com/admin/users/" + str(user_num)
    print(user_link)

    browser.get(user_link)
    time.sleep(1.5)

    # إدخال الأرباح
    user_rev_unformatted = ws["T" + str(row)].value
    user_rev = "{:.2f}".format(user_rev_unformatted)
    print(user_rev)
    browser.find_element(
        By.CSS_SELECTOR,
        "body > div > div > section.content > div:nth-child(4) > div > div.box-body > form > input.form-control",
    ).click()
    time.sleep(0.75)
    browser.find_element(
        By.CSS_SELECTOR,
        "body > div > div > section.content > div:nth-child(4) > div > div.box-body > form > input.form-control",
    ).send_keys(user_rev)
    time.sleep(0.75)

    browser.find_element(
        By.CSS_SELECTOR,
        "body > div > div > section.content > div:nth-child(4) > div > div.box-body > form > button",
    ).click()

    # تحويل إلى مستحقات متراكمة
    browser.find_element(
        By.CSS_SELECTOR,
        "body > div > div > section.content > div:nth-child(11) > div > div.box-body > form > button",
    ).click()
